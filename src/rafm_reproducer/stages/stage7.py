"""
Stage 7 — MECHANICAL: generate the Excel workbook from the Stage 5 spec
and Stage 6 omissions via openpyxl.

Sheet layout:
  Doc    — verbatim source code of every selected formula
  Inputs — scalar parameters (named ranges) + scenario table
  Model  — time-series projection with live Excel formulas
  Notes  — scope notes + Stage 6 omissions
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from ..artifacts import save_json
from ..schemas.pipeline_state import EnrichedFormula, PipelineState
from ..schemas.stage5 import Stage5Output
from ..schemas.stage6 import Stage6Output

# ── colour palette ─────────────────────────────────────────────────────────
_BLUE_HEADER = PatternFill("solid", fgColor="2F5597")
_YELLOW_INPUT = PatternFill("solid", fgColor="FFF2CC")
_GREEN_HEADER = PatternFill("solid", fgColor="548235")
_GREY_ROW = PatternFill("solid", fgColor="F2F2F2")
_WHITE = PatternFill("solid", fgColor="FFFFFF")

_H_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)


# ── helpers ────────────────────────────────────────────────────────────────

def _hdr(ws, row: int, col: int, value: str, fill=_BLUE_HEADER):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    c.font = _H_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def _val(ws, row: int, col: int, value, fill=_WHITE, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill
    if bold:
        c.font = _BOLD
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def _resolve_formula(
    formula_excel: str,
    col_map: dict[str, str],
    current_row: int,
    header_row: int,
    input_names: set[str],
) -> str:
    """
    Replace column_name references with absolute Excel cell references.
    - prev_X / X_prev → column X, previous row (or header row for t=1)
    - regular X (not an input) → column X, current row
    - input names → left as-is (they are Excel named ranges)
    """
    result = formula_excel
    prev_row = current_row - 1 if current_row > header_row + 1 else header_row

    # Sort by length desc to avoid partial replacements
    ordered = sorted(col_map.items(), key=lambda kv: len(kv[0]), reverse=True)

    for col_name, col_letter in ordered:
        prev_ref = f"{col_letter}{prev_row}"
        cur_ref = f"{col_letter}{current_row}"

        # prev_X
        result = re.sub(r'\bprev_' + re.escape(col_name) + r'\b', prev_ref, result)
        # X_prev
        result = re.sub(r'\b' + re.escape(col_name) + r'_prev\b', prev_ref, result)
        # regular X (only for calculated columns, not inputs)
        if col_name not in input_names:
            result = re.sub(r'\b' + re.escape(col_name) + r'\b', cur_ref, result)

    prefix = "=" if not result.startswith("=") else ""
    return prefix + result


# ── Sheet builders ─────────────────────────────────────────────────────────

def _build_doc_sheet(ws, enriched: list[EnrichedFormula]):
    ws.title = "Doc"
    ws.sheet_view.showGridLines = True

    headers = ["Numero", "Titre", "Page", "Source Code"]
    widths = [14, 30, 8, 100]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        _hdr(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = w

    for r, ef in enumerate(enriched, 2):
        fill = _GREY_ROW if r % 2 == 0 else _WHITE
        _val(ws, r, 1, ef.numero, fill)
        _val(ws, r, 2, ef.titre, fill)
        _val(ws, r, 3, ef.page, fill)
        c = _val(ws, r, 4, ef.contenu, fill)
        c.alignment = Alignment(wrap_text=False, vertical="top")
        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "A2"


def _build_inputs_sheet(ws, spec: Stage5Output, wb: Workbook):
    ws.title = "Inputs"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18

    # Section header
    _hdr(ws, 1, 1, "Parameter", fill=_GREEN_HEADER)
    _hdr(ws, 1, 2, "Label", fill=_GREEN_HEADER)
    _hdr(ws, 1, 3, "Value", fill=_GREEN_HEADER)

    for r, inp in enumerate(spec.inputs, 2):
        _val(ws, r, 1, inp.name)
        _val(ws, r, 2, inp.label)
        cell = ws.cell(row=r, column=3, value=inp.default_value)
        cell.fill = _YELLOW_INPUT
        if inp.comment:
            ws.cell(row=r, column=4, value=f"# {inp.comment}")

        # Named range for this input (points to the value cell in this sheet)
        col_letter = get_column_letter(3)
        ref = f"Inputs!${col_letter}${r}"
        safe_name = re.sub(r'[^A-Za-z0-9_]', '_', inp.name)
        try:
            wb.defined_names.add(DefinedName(safe_name, attr_text=ref))
        except Exception:
            pass  # duplicate or invalid name — skip silently

    # Scenario section
    scenario_start = len(spec.inputs) + 3
    driving = spec.scenario.driving_input
    _hdr(ws, scenario_start, 1, "Scenario", fill=_BLUE_HEADER)
    _hdr(ws, scenario_start, 2, f"{driving} (scenario)", fill=_BLUE_HEADER)

    for t, val in enumerate(spec.scenario.default_values, 1):
        row = scenario_start + t
        _val(ws, row, 1, t)
        cell = ws.cell(row=row, column=2, value=val)
        cell.fill = _YELLOW_INPUT

    ws.freeze_panes = "A2"


def _build_model_sheet(ws, spec: Stage5Output):
    ws.title = "Model"

    calcs = spec.calculations
    input_names = {inp.name for inp in spec.inputs}
    n = spec.scenario.n_periods

    # Column layout: A=Period, B onwards = calculations
    HEADER_ROW = 1
    INIT_ROW = 2    # initial-values row (all zeros / empty) — prev refs from t=1 land here
    DATA_START = 3  # t=1

    # Column map: column_name → Excel column letter
    col_map: dict[str, str] = {}
    for i, calc in enumerate(calcs):
        col_map[calc.column_name] = get_column_letter(i + 2)  # B, C, D ...

    # Set column widths
    ws.column_dimensions["A"].width = 10
    for col_name, col_letter in col_map.items():
        ws.column_dimensions[col_letter].width = 16

    # Header row
    _hdr(ws, HEADER_ROW, 1, "Period")
    for calc, col_letter in zip(calcs, col_map.values()):
        hdr_cell = _hdr(ws, HEADER_ROW, ord(col_letter) - ord("A") + 1, calc.label)
        hdr_cell.alignment = Alignment(wrap_text=True, horizontal="center")

    # Init row (row 2) — zeros for calculated columns
    ws.cell(row=INIT_ROW, column=1, value=0).font = Font(color="AAAAAA")
    for col_letter in col_map.values():
        c = ws.cell(row=INIT_ROW, column=ord(col_letter) - ord("A") + 1, value=0)
        c.font = Font(color="AAAAAA")

    # Data rows (t=1 to n)
    for t in range(1, n + 1):
        current_row = DATA_START + t - 1
        fill = _GREY_ROW if t % 2 == 0 else _WHITE

        ws.cell(row=current_row, column=1, value=t).fill = fill

        for calc in calcs:
            col_letter = col_map[calc.column_name]
            col_idx = ord(col_letter) - ord("A") + 1

            if calc.kind == "conditional" and calc.condition_excel:
                condition = _resolve_formula(
                    calc.condition_excel, col_map, current_row, HEADER_ROW, input_names
                ).lstrip("=")
                body = _resolve_formula(
                    calc.formula_excel, col_map, current_row, HEADER_ROW, input_names
                ).lstrip("=")
                formula = f"=IF({condition},{body},0)"
            else:
                formula = _resolve_formula(
                    calc.formula_excel, col_map, current_row, HEADER_ROW, input_names
                )

            cell = ws.cell(row=current_row, column=col_idx, value=formula)
            cell.fill = fill
            if calc.format in ("percent", "pct", "%"):
                cell.number_format = "0.00%"
            elif calc.format in ("number", "int", "integer"):
                cell.number_format = "#,##0"
            elif calc.format in ("decimal", "float"):
                cell.number_format = "#,##0.000"

    ws.freeze_panes = f"B{DATA_START}"


def _build_notes_sheet(ws, spec: Stage5Output, stage6: Stage6Output | None):
    ws.title = "Notes"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90

    row = 1
    _hdr(ws, row, 1, "Section")
    _hdr(ws, row, 2, "Content")
    row += 1

    # Mechanism overview
    _val(ws, row, 1, "Mechanism", bold=True)
    _val(ws, row, 2, spec.mechanism_name)
    row += 1
    _val(ws, row, 1, "Scope")
    _val(ws, row, 2, spec.scope_description)
    row += 1
    if spec.key_relation:
        _val(ws, row, 1, "Key identity")
        _val(ws, row, 2, spec.key_relation.get("identity", ""))
        row += 1

    row += 1  # blank

    # Notes from Stage 5
    for note in spec.notes:
        _val(ws, row, 1, note.section, bold=True)
        c = _val(ws, row, 2, note.text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60
        row += 1

    row += 1  # blank

    # Omissions from Stage 6
    if stage6 and stage6.omissions:
        _hdr(ws, row, 1, "Omissions & simplifications")
        _hdr(ws, row, 2, "Impact")
        row += 1

        for omission in sorted(stage6.omissions, key=lambda o: o.kind):
            kind_label = {"concerning": "⚠️ Concerning", "branch": "Branch", "deliberate": "Deliberate"}.get(
                omission.kind, omission.kind
            )
            _val(ws, row, 1, f"[{kind_label}] {omission.summary}", bold=(omission.kind == "concerning"))
            _val(ws, row, 2, f"{omission.impact_one_line}\nSource: [{omission.source_formula_number}] — {omission.source_excerpt}")
            ws.row_dimensions[row].height = 50
            row += 1

    if stage6 and stage6.overall_concerns:
        row += 1
        _hdr(ws, row, 1, "Overall concerns")
        row += 1
        for concern in stage6.overall_concerns:
            _val(ws, row, 2, concern)
            row += 1


# ── Main entry ─────────────────────────────────────────────────────────────

def run_stage7(state: PipelineState) -> PipelineState:
    assert state.stage5_output is not None, "Stage 5 must complete before Stage 7"
    assert state.checkpoint2 is not None and state.checkpoint2.action == "approve"

    spec = state.stage5_output
    stage6 = state.stage6_output
    enriched = state.stage3_enriched

    wb = Workbook()
    # Remove the default blank sheet
    if wb.active:
        wb.remove(wb.active)

    ws_doc = wb.create_sheet("Doc")
    ws_inputs = wb.create_sheet("Inputs")
    ws_model = wb.create_sheet("Model")
    ws_notes = wb.create_sheet("Notes")

    _build_doc_sheet(ws_doc, enriched)
    _build_inputs_sheet(ws_inputs, spec, wb)
    _build_model_sheet(ws_model, spec)
    _build_notes_sheet(ws_notes, spec, stage6)

    output_path = state.run_dir / "output.xlsx"
    wb.save(str(output_path))

    state.output_xlsx_path = output_path
    save_json(state.run_dir, "pipeline_state", state)
    return state
